import uuid
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.config import settings


def generate_professional_xlsx(
    title: str,
    headers: list[str],
    data_rows: list[list],
    summary_stats: dict | None = None,
    file_prefix: str = "Industrial_Data"
) -> str:
    """Generates a styled Excel workbook with summary sheet, data sheet, and formatting."""
    doc_id = uuid.uuid4().hex[:6].upper()
    file_name = f"{file_prefix}_{datetime.utcnow().strftime('%Y%m%d')}_{doc_id}.xlsx"
    file_path = settings.ARTIFACTS_DIR / file_name

    wb = openpyxl.Workbook()
    
    # Sheet 1: Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    
    ws_summary["A1"] = "MRPL SOVEREIGN INDUSTRIAL DATA REPORT"
    ws_summary["A1"].font = Font(name="Arial", size=14, bold=True, color="1E3A8A")
    
    ws_summary["A3"] = "Report Title:"
    ws_summary["B3"] = title
    ws_summary["A3"].font = Font(bold=True)
    
    ws_summary["A4"] = "Generated At:"
    ws_summary["B4"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    ws_summary["A4"].font = Font(bold=True)

    if summary_stats:
        ws_summary["A6"] = "Key Summary Statistics"
        ws_summary["A6"].font = Font(size=12, bold=True, color="1E3A8A")
        r = 7
        for k, v in summary_stats.items():
            ws_summary.cell(row=r, column=1, value=str(k)).font = Font(bold=True)
            ws_summary.cell(row=r, column=2, value=str(v))
            r += 1

    # Sheet 2: Data Sheet
    ws_data = wb.create_sheet(title="Detailed Records")
    
    # Header styling
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")

    # Auto-fit column widths
    for sheet in [ws_summary, ws_data]:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(str(file_path))
    return str(file_path)
